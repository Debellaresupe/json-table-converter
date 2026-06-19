from __future__ import annotations

from io import BytesIO
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

Path = tuple[str, ...]


def build_hierarchical_table(data: Any, max_depth: int = 80, preview_limit: int | None = None) -> dict[str, Any]:
    columns = _collect_leaf_paths(data, max_depth=max_depth)
    if not columns:
        columns = [("value",)]

    rows = _render_rows(data, tuple(), set(columns), 0, max_depth)
    if not rows:
        rows = [{}]

    total_rows = len(rows)
    if preview_limit is not None:
        rows = rows[:preview_limit]

    header_depth = max(len(c) for c in columns)
    header_rows = _build_header_rows(columns, header_depth)
    data_rows = [
        [_excel_cell(row.get(path)) for path in columns]
        for row in rows
    ]

    return {
        "header_rows": header_rows,
        "rows": data_rows,
        "meta": {
            "row_count": total_rows,
            "column_count": len(columns),
            "returned_rows": len(rows),
            "truncated": preview_limit is not None and total_rows > len(rows),
        },
    }


def to_hierarchical_xlsx_bytes(data: Any, max_depth: int = 80) -> bytes:
    table = build_hierarchical_table(data, max_depth=max_depth)
    header_rows: list[list[Any]] = table["header_rows"]
    rows: list[list[Any]] = table["rows"]
    header_depth = len(header_rows)
    col_count = len(header_rows[0]) if header_rows else 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for ridx, row in enumerate(header_rows, start=1):
        for cidx, value in enumerate(row, start=1):
            ws.cell(row=ridx, column=cidx, value=value)

    for ridx, row in enumerate(rows, start=header_depth + 1):
        for cidx, value in enumerate(row, start=1):
            ws.cell(row=ridx, column=cidx, value=value)

    ws.freeze_panes = ws.cell(row=header_depth + 1, column=1).coordinate
    _style_sheet(ws, header_depth, len(rows), col_count)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_header_rows(columns: list[Path], header_depth: int) -> list[list[Any]]:
    """Build a sparse multi-row header exactly like the reference XLSX.

    Repeated parent path labels are written only once at the beginning of each
    group. Example: [shop.id, shop.name] becomes row1: [shop, blank],
    row2: [id, name].
    """
    header_rows: list[list[Any]] = []
    for level in range(header_depth):
        header_row: list[Any] = []
        previous_prefix: Path | None = None
        for path in columns:
            if level >= len(path):
                header_row.append(None)
                continue
            current_prefix = path[: level + 1]
            header_row.append(path[level] if current_prefix != previous_prefix else None)
            previous_prefix = current_prefix
        header_rows.append(header_row)
    return header_rows


def _collect_leaf_paths(value: Any, prefix: Path = tuple(), max_depth: int = 80, depth: int = 0) -> list[Path]:
    if depth > max_depth:
        return [prefix or ("value",)]
    if _is_primitive(value):
        return [prefix or ("value",)]
    if isinstance(value, dict):
        out: list[Path] = []
        if not value:
            return [prefix or ("value",)]
        for key, child in value.items():
            out.extend(_collect_leaf_paths(child, prefix + (str(key),), max_depth, depth + 1))
        return _dedupe(out)
    if isinstance(value, list):
        if not value:
            return [prefix or ("value",)]
        out: list[Path] = []
        for item in value:
            out.extend(_collect_leaf_paths(item, prefix, max_depth, depth + 1))
        return _dedupe(out)
    return [prefix or ("value",)]


def _render_rows(value: Any, prefix: Path, leaf_paths: set[Path], depth: int, max_depth: int) -> list[dict[Path, Any]]:
    if depth > max_depth:
        return [{prefix or ("value",): _json(value)}]
    if _is_primitive(value):
        return [{prefix or ("value",): value}]
    if isinstance(value, dict):
        if not value:
            return [{prefix or ("value",): "{}"}]
        child_blocks: list[list[dict[Path, Any]]] = []
        for key, child in value.items():
            child_blocks.append(_render_rows(child, prefix + (str(key),), leaf_paths, depth + 1, max_depth))
        return _align_blocks(child_blocks)
    if isinstance(value, list):
        if not value:
            return []
        out: list[dict[Path, Any]] = []
        for item in value:
            out.extend(_render_rows(item, prefix, leaf_paths, depth + 1, max_depth))
        return out
    return [{prefix or ("value",): str(value)}]


def _align_blocks(blocks: list[list[dict[Path, Any]]]) -> list[dict[Path, Any]]:
    row_count = max((len(block) for block in blocks), default=1)
    rows: list[dict[Path, Any]] = []
    for idx in range(row_count):
        merged: dict[Path, Any] = {}
        for block in blocks:
            if idx < len(block):
                merged.update(block[idx])
        rows.append(merged)
    return rows


def _style_sheet(ws, header_depth: int, data_rows: int, col_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="EAF1F8")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=header_depth, min_col=1, max_col=col_count):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    for row in ws.iter_rows(min_row=header_depth + 1, max_row=header_depth + data_rows, min_col=1, max_col=col_count):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = border
    for idx in range(1, col_count + 1):
        width = 10
        for col_cells in ws.iter_cols(min_col=idx, max_col=idx, min_row=1, max_row=min(header_depth + data_rows, header_depth + 80)):
            for cell in col_cells:
                width = max(width, min(len(str(cell.value or "")) + 2, 42))
        ws.column_dimensions[get_column_letter(idx)].width = width


def _is_primitive(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def _dedupe(paths: list[Path]) -> list[Path]:
    seen: set[Path] = set()
    out: list[Path] = []
    for path in paths:
        if path not in seen:
            seen.add(path)
            out.append(path)
    return out


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _excel_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return _json(value)
