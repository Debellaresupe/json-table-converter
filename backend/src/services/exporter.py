from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def ordered_columns(rows: list[dict[str, Any]], columns: list[str] | None = None) -> list[str]:
    if columns:
        return columns
    seen: set[str] = set()
    out: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                out.append(key)
    return out


def to_csv_bytes(rows: list[dict[str, Any]], columns: list[str] | None, delimiter: str = ",") -> bytes:
    cols = ordered_columns(rows, columns)
    buffer = StringIO(newline="")
    buffer.write("\ufeff")
    writer = csv.DictWriter(buffer, fieldnames=cols, delimiter=delimiter[:1] or ",", extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({c: _cell(row.get(c)) for c in cols})
    return buffer.getvalue().encode("utf-8")


def to_xlsx_bytes(rows: list[dict[str, Any]], columns: list[str] | None) -> bytes:
    cols = ordered_columns(rows, columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "JSON Table"
    ws.append(cols)
    for row in rows:
        ws.append([_cell(row.get(c)) for c in cols])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, col in enumerate(cols, 1):
        max_len = min(max([len(str(col))] + [len(str(r.get(col, ""))) for r in rows[:500]]), 60)
        ws.column_dimensions[get_column_letter(idx)].width = max(10, max_len + 2)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
